from odoo import fields, models
import base64
import openpyxl
import io

import datetime


class dataUpload(models.TransientModel):
    _name = 'data.upload'

    data_file = fields.Binary(string='File', required=True)

    def import_data(self):

        decoded_data = base64.b64decode(self.data_file)
        xls_filelike = io.BytesIO(decoded_data)
        workbook = openpyxl.load_workbook(xls_filelike)
        sheet = workbook.active
        row = 1 + sheet.min_row
        mylist = []
        partner_obj = self.env["res.partner"]
        partner_name = sheet.cell(row, 1).value
        partner = partner_obj.search([('name', '=ilike', partner_name)], limit=1)
        nrows = sheet.max_row
        purchase_order = self.env['purchase.order']
        while row <= nrows + 1:
            if sheet.cell(row, 1).value != partner_name:
                values = {
                    'partner_id': partner.id,
                    'order_line': mylist
                }

                purchase_order = purchase_order.create(values)
                if row == nrows + 1:
                    break
                partner_name = sheet.cell(row, 1).value
                partner = partner_obj.search([('name', '=ilike', partner_name)], limit=1)
                mylist = []

            product = self.env['product.product'].search([('default_code', '=ilike', sheet.cell(row, 2).value)])
            qty = str(sheet.cell(row, 3).value)
            rate = str(sheet.cell(row, 4).value)
            date = datetime.datetime.today()
            vals = (0, 0, {'product_id': product.id, 'name': product.name, 'product_qty': qty, 'price_unit': rate,
                           'date_planned': date, 'product_uom': product.uom_id.id})

            if product:
                mylist.append(vals)
            else:
                print("Product not found on line # ", row)
            row += 1

        purchase_order.button_confirm()
        for pick in purchase_order.picking_ids:
            pick.action_confirm()

            pick.action_assign()

            from odoo.tests import Form
            wiz_act = pick.button_validate()
            wiz = Form(self.env[wiz_act['res_model']].with_context(wiz_act['context'])).save()
            wiz.process()
