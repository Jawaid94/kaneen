# -*- coding: iso-8859-1 -*-

import openpyxl
import base64
import io
import datetime
import logging

from odoo import models, fields, api

_logger = logging.getLogger(__name__)


class ImportProduct(models.TransientModel):
    _name = 'import.product'

    sale_tax_id = fields.Many2many('account.tax', 'account_tax_sale_tax_rel', domain=[('type_tax_use', '=', 'sale')])
    purchase_tax_id = fields.Many2many('account.tax', 'account_tax_purchase_tax_rel',
                                       domain=[('type_tax_use', '=', 'purchase')])
    pricelist_id = fields.Many2one('product.pricelist')
    start_date_for_price = fields.Date()
    file = fields.Binary()

    def import_file(self):
        decrypted = base64.b64decode(self.file)
        xls_file_like = io.BytesIO(decrypted)
        workbook = openpyxl.load_workbook(xls_file_like)
        sheet_obj = workbook.active
        max_row = sheet_obj.max_row
        broke_barcode = []
        for i in range(2, max_row + 1):
            barcode = sheet_obj.cell(row=i, column=1).value
            default_code = sheet_obj.cell(row=i, column=2).value
            name = sheet_obj.cell(row=i, column=6).value
            seller_name = sheet_obj.cell(row=i, column=7).value
            price = sheet_obj.cell(row=i, column=22).value
            list_price = sheet_obj.cell(row=i, column=23).value
            name_arab = sheet_obj.cell(row=i, column=5).value
            size = sheet_obj.cell(row=i, column=21).value
            size_arab = sheet_obj.cell(row=i, column=20).value
            if not barcode:
                break
            product_id = self.env['product.template'].search([('default_code', '=', default_code)], limit=1)
            route_ids = product_id.route_ids.search([]).filtered(
                lambda x: x.name == 'Buy' or x.name == 'Make To Order + Make To Stock')
            if not product_id:
                product_id = self.env['product.template'].create({
                    'name': name + '-' + str(size),
                    'default_code': default_code,
                    # 'barcode': barcode,
                    'sale_ok': True,
                    'purchase_ok': True,
                    'route_ids': [(6, 0, route_ids.ids)],
                    'list_price': list_price,
                    'supplier_taxes_id': [(6, 0, self.purchase_tax_id.ids)],
                    'taxes_id': [(6, 0, self.sale_tax_id.ids)],
                    'detailed_type': 'product'
                })
            else:
                product_id.write({
                    'name': name + '-' + str(size),
                    'sale_ok': True,
                    'purchase_ok': True,
                    'list_price': list_price,
                    'detailed_type': 'product'
                    # 'route_ids': [(6, 0, route_ids.ids)]
                })
                pricelist_item = self.pricelist_id.item_ids.search(
                    [('pricelist_id', '=', self.pricelist_id.id), ('product_tmpl_id', '=', product_id.id)])
                if pricelist_item:
                    pricelist_item.write({
                        'date_end': datetime.datetime(self.start_date_for_price.year, self.start_date_for_price.month,
                                                      self.start_date_for_price.day, 23, 59, 59, )
                    })
            self.pricelist_id.item_ids.create({
                'name': product_id.name,
                'fixed_price': list_price,
                'product_tmpl_id': product_id.id,
                'applied_on': '1_product',
                'pricelist_id': self.pricelist_id.id,
                'date_start': datetime.datetime(self.start_date_for_price.year, self.start_date_for_price.month,
                                                self.start_date_for_price.day, 0, 0, 0, )
            })

            translation = self.env['ir.translation'].search(
                [('res_id', '=', product_id.id), ('lang', '=', 'ar_001'), ('state', '=', 'translated'),
                 ('name', '=', 'product.template,name')])
            if translation:
                translation.write({
                    'value': name_arab + '-' + str(size_arab)
                })
            else:
                translation = self.env['ir.translation'].create({
                    'type': 'model',
                    'name': 'product.template,name',
                    'lang': 'ar_001',
                    'res_id': product_id.id,
                    'value': name_arab + '-' + str(size_arab),
                    'state': 'translated',
                })
                self.env['ir.translation'].translate_fields('product.template', product_id.id, 'name')
            product_id.write({'tranlate_field': name_arab})
            barcode_check = product_id.search([('barcode', '=', str(barcode))])
            barcode_ids_check = product_id.barcode_ids.search([('name', '=', str(barcode))])
            create_barcode = True
            if barcode_check and barcode_check.id != product_id.id:
                broke_barcode.append(str(barcode))
                create_barcode = False
            if barcode_ids_check and barcode_ids_check.product_id.id != product_id.product_variant_id.id:
                broke_barcode.append(str(barcode))
                create_barcode = False
            barcode_ids = product_id.barcode_ids.filtered(lambda x: x.name == str(barcode))
            if str(barcode) != product_id.barcode and not barcode_ids and create_barcode:
                if not product_id.barcode:
                    product_id.write({
                        'barcode': barcode
                    })
                else:
                    barcode_ids.create({
                        'name': barcode,
                        'product_id': product_id.product_variant_id.id
                    })

            partner_seller_id = self.env['res.partner'].search([('name', '=', seller_name)])
            if not partner_seller_id:
                partner_seller_id = partner_seller_id.create({
                    'name': seller_name
                })
                seller_id = self.env['product.supplierinfo'].create({
                    'name': partner_seller_id.id,
                    'product_tmpl_id': product_id.id
                })
            else:
                seller_id = self.env['product.supplierinfo'].search(
                    [('name', '=', partner_seller_id.id), ('product_tmpl_id', '=', product_id.id)])
                if not seller_id:
                    seller_id = self.env['product.supplierinfo'].create({
                        'name': partner_seller_id.id,
                        'product_tmpl_id': product_id.id
                    })
            seller_ids = product_id.seller_ids.filtered(lambda x: x.id == seller_id.id)
            if seller_ids:
                seller_ids.write({
                    'currency_id': partner_seller_id.currency_id.id,
                    'price': price
                })
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'type': 'warning',
                'message': "Barcode {}  is not unique".format(' '.join(set(broke_barcode))),
                'sticky': True,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }

    def import_file_new(self, local=False):
        if local:
            workbook = openpyxl.load_workbook("/opt/odoo/kaneen/mo_kaneen/wizard/products.xlsx")
        else:
            decrypted = base64.b64decode(self.file)
            xls_file_like = io.BytesIO(decrypted)
            workbook = openpyxl.load_workbook(xls_file_like)
        sheet_obj = workbook.active
        max_row = sheet_obj.max_row
        broke_barcode = []
        for i in range(2, max_row + 1):
            _logger.error(i)
            additional_barcode = []
            sku_shatha = str(sheet_obj.cell(row=i, column=1).value).replace('.0', '')
            default_code = sheet_obj.cell(row=i, column=2).value
            name_arab = sheet_obj.cell(row=i, column=3).value
            name = sheet_obj.cell(row=i, column=4).value
            vendor = sheet_obj.cell(row=i, column=5).value
            vendor_price = sheet_obj.cell(row=i, column=6).value
            pricelst_price = sheet_obj.cell(row=i, column=7).value
            barcode = str(sheet_obj.cell(row=i, column=9).value).replace('.0', '')
            for k in range(1, 24):
                ad_barcode = str(sheet_obj.cell(row=i, column=9 + k).value).replace('.0', '')
                if ad_barcode == 'None':
                    break
                else:
                    additional_barcode.append(ad_barcode)

            product_id = self.env['product.template'].search([('default_code', '=', default_code)], limit=1)
            route_ids = product_id.route_ids.search([]).filtered(
                lambda x: x.name == 'Buy' or x.name == 'Make To Order + Make To Stock')
            if not product_id:
                try:
                    product_id = self.env['product.template'].create({
                        'name': name,
                        'default_code': default_code,
                        'barcode': barcode,
                        'sku_shatha': sku_shatha,
                        'sale_ok': True,
                        'purchase_ok': True,
                        'route_ids': [(6, 0, route_ids.ids)],
                        'list_price': pricelst_price,
                        'supplier_taxes_id': [(6, 0, self.purchase_tax_id.ids)],
                        'taxes_id': [(6, 0, self.sale_tax_id.ids)],
                        'detailed_type': 'product'
                    })
                except Exception:
                    broke_barcode.append(barcode)
                    continue
                for code in additional_barcode:
                    ad_barcode = self.env['product.barcode.multi'].search(
                        [('name', '=', code), ('product_id', '=', product_id.product_variant_id.id)])
                    if not ad_barcode:
                        ad_barcode = ad_barcode.create({
                            'name': code,
                            'product_id': product_id.product_variant_id.id
                        })

            else:
                product_id.write({
                    'name': name,
                    'sale_ok': True,
                    'purchase_ok': True,
                    'list_price': pricelst_price,
                    'detailed_type': 'product'
                })
                pricelist_item = self.pricelist_id.item_ids.search(
                    [('pricelist_id', '=', self.pricelist_id.id), ('product_tmpl_id', '=', product_id.id)])
                if pricelist_item:
                    pricelist_item.write({
                        'date_end': datetime.datetime(self.start_date_for_price.year, self.start_date_for_price.month,
                                                      self.start_date_for_price.day, 23, 59, 59, )
                    })
            self.pricelist_id.item_ids.create({
                'name': product_id.name,
                'fixed_price': pricelst_price,
                'product_tmpl_id': product_id.id,
                'applied_on': '1_product',
                'pricelist_id': self.pricelist_id.id,
                'date_start': datetime.datetime(self.start_date_for_price.year, self.start_date_for_price.month,
                                                self.start_date_for_price.day, 0, 0, 0, )
            })

            translation = self.env['ir.translation'].search(
                [('res_id', '=', product_id.id), ('lang', '=', 'ar_001'), ('state', '=', 'translated'),
                 ('name', '=', 'product.template,name')])
            if translation:
                translation.write({
                    'value': name_arab
                })
            else:
                translation = self.env['ir.translation'].create({
                    'type': 'model',
                    'name': 'product.template,name',
                    'lang': 'ar_001',
                    'res_id': product_id.id,
                    'value': name_arab,
                    'state': 'translated',
                })
                self.env['ir.translation'].translate_fields('product.template', product_id.id, 'name')
            product_id.write({'tranlate_field': name_arab})

            # barcode_check = product_id.search([('barcode', '=', str(barcode))])
            # barcode_ids_check = product_id.barcode_ids.search([('name', '=', str(barcode))])
            # create_barcode = True
            # if barcode_check and barcode_check.id != product_id.id:
            #     broke_barcode.append(str(barcode))
            #     create_barcode = False
            # if barcode_ids_check and barcode_ids_check.product_id.id != product_id.product_variant_id.id:
            #     broke_barcode.append(str(barcode))
            #     create_barcode = False
            # barcode_ids = product_id.barcode_ids.filtered(lambda x: x.name == str(barcode))
            # if str(barcode) != product_id.barcode and not barcode_ids and create_barcode:
            #     if not product_id.barcode:
            #         product_id.write({
            #             'barcode': barcode
            #         })
            #     else:
            #         barcode_ids.create({
            #             'name': barcode,
            #             'product_id': product_id.product_variant_id.id
            #         })

            partner_seller_id = self.env['res.partner'].search([('name', '=', vendor)])
            if not partner_seller_id:
                partner_seller_id = partner_seller_id.create({
                    'name': vendor
                })
                seller_id = self.env['product.supplierinfo'].create({
                    'name': partner_seller_id.id,
                    'product_tmpl_id': product_id.id
                })
            else:
                seller_id = self.env['product.supplierinfo'].search(
                    [('name', '=', partner_seller_id.id), ('product_tmpl_id', '=', product_id.id)])
                if not seller_id:
                    seller_id = self.env['product.supplierinfo'].create({
                        'name': partner_seller_id.id,
                        'product_tmpl_id': product_id.id
                    })
            seller_ids = product_id.seller_ids.filtered(lambda x: x.id == seller_id.id)
            if seller_ids:
                seller_ids.write({
                    'currency_id': partner_seller_id.currency_id.id,
                    'price': vendor_price
                })
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'type': 'warning',
                'message': "Barcode {}  is not unique".format(' '.join(set(broke_barcode))),
                'sticky': True,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }
